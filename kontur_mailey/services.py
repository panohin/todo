import openpyxl 


def copy_sheet():
	wb_target = openpyxl.Workbook()
	target_sheet = wb_target.create_sheet('123')

def handle_uploaded_file(file):
	s = openpyxl.load_workbook(file)
	print(dir(s))
	for sheet_name in s.get_sheet_names():
		sheet = s.get_sheet_by_name(sheet_name)
		print(type(sheet))
	